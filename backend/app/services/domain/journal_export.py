from __future__ import annotations

import re
from datetime import UTC, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.cell import TYPE_STRING
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.security import transliterate
from app.models import AssignmentStatus, StudyGroup, WorkKind
from app.services.domain.journal import Journal
from app.services.domain.xlsx import text_cell

NAME_COLUMN_WIDTH = 32
GRADE_COLUMN_WIDTH = 9
PERCENT_COLUMN_WIDTH = 7
AVERAGE_COLUMN_WIDTH = 12

HEADER_FILL = PatternFill("solid", fgColor="EEF2FF")
TOTAL_FILL = PatternFill("solid", fgColor="F5F5F4")
THIN = Side(style="thin", color="D4D4D8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Excel запрещает эти символы в имени листа и ограничивает его 31 знаком.
_SHEET_FORBIDDEN = re.compile(r"[\[\]:*?/\\]")
_SHEET_LIMIT = 31

KIND_LABELS: dict[WorkKind, str] = {
    WorkKind.HOMEWORK: "Домашние задания",
    WorkKind.CONTROL: "Контрольные работы",
}


def _sheet_title(group: StudyGroup) -> str:
    cleaned = _SHEET_FORBIDDEN.sub(" ", group.name).strip()
    return (cleaned or "Журнал")[:_SHEET_LIMIT]


def _status_note(status: AssignmentStatus) -> str:
    return {
        AssignmentStatus.NEW: "не начата",
        AssignmentStatus.IN_PROGRESS: "выполняется",
        AssignmentStatus.SUBMITTED: "",
        AssignmentStatus.OVERDUE: "просрочена",
    }.get(status, "")


def build_workbook(
    journal: Journal,
    *,
    group: StudyGroup,
    filters: dict[str, str | None] | None = None,
    generated_at: datetime | None = None,
) -> bytes:
    moment = generated_at or datetime.now(UTC)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_title(group)

    # шапка: что именно выгружено
    applied = [value for value in (filters or {}).values() if value]
    text_cell(
        sheet, row=1, column=1, value=f"Журнал группы «{group.name}»"
    ).font = Font(bold=True, size=13)
    subtitle = f"Выгружено {moment.astimezone().strftime('%d.%m.%Y %H:%M')}"
    if applied:
        subtitle = f"{subtitle} · фильтры: {', '.join(applied)}"
    text_cell(sheet, row=2, column=1, value=subtitle).font = Font(
        size=10, color="52525B"
    )

    header_top = 4
    header_bottom = 5
    first_data_row = 6

    sheet.cell(row=header_top, column=1, value="Фамилия и имя")
    sheet.merge_cells(
        start_row=header_top, start_column=1, end_row=header_bottom, end_column=1
    )

    # Отметка и процент — двумя колонками под общим заголовком работы: в одной
    # ячейке они остались бы текстом, раздельно их можно сортировать и считать.
    column = 2
    for item in journal.columns:
        text_cell(
            sheet, row=header_top, column=column, value=item.work.display_title
        )
        sheet.merge_cells(
            start_row=header_top,
            start_column=column,
            end_row=header_top,
            end_column=column + 1,
        )
        sheet.cell(row=header_bottom, column=column, value="Оценка")
        sheet.cell(row=header_bottom, column=column + 1, value="%")
        sheet.column_dimensions[get_column_letter(column)].width = GRADE_COLUMN_WIDTH
        sheet.column_dimensions[get_column_letter(column + 1)].width = (
            PERCENT_COLUMN_WIDTH
        )
        column += 2

    totals_start = column
    for title in (
        "Средняя",
        "Средняя ДЗ",
        "Средняя КР",
        "Итог за четверть",
    ):
        sheet.cell(row=header_top, column=column, value=title)
        sheet.merge_cells(
            start_row=header_top, start_column=column, end_row=header_bottom,
            end_column=column,
        )
        sheet.column_dimensions[get_column_letter(column)].width = (
            AVERAGE_COLUMN_WIDTH
        )
        column += 1
    last_column = column - 1

    for row in (header_top, header_bottom):
        for index in range(1, last_column + 1):
            cell = sheet.cell(row=row, column=index)
            cell.font = Font(bold=True, size=10)
            cell.alignment = CENTER
            cell.fill = HEADER_FILL
            cell.border = BORDER

    # строки учеников
    for offset, entry in enumerate(journal.rows):
        row_index = first_data_row + offset
        name_cell = text_cell(
            sheet, row=row_index, column=1, value=entry.student.full_name
        )
        name_cell.alignment = Alignment(horizontal="left", vertical="center")
        name_cell.border = BORDER

        column = 2
        for item in journal.columns:
            journal_cell = entry.cells.get(item.work.id)
            grade_cell = sheet.cell(row=row_index, column=column)
            percent_cell = sheet.cell(row=row_index, column=column + 1)
            if journal_cell is not None:
                if journal_cell.grade is not None:
                    grade_cell.value = journal_cell.grade
                else:
                    grade_cell.value = _status_note(journal_cell.status)
                    grade_cell.data_type = TYPE_STRING
                    grade_cell.font = Font(size=9, color="A1A1AA")
                if journal_cell.percent is not None:
                    percent_cell.value = round(journal_cell.percent)
                    percent_cell.number_format = '0"%"'
            for target in (grade_cell, percent_cell):
                target.alignment = CENTER
                target.border = BORDER
            column += 2

        values = (
            entry.average_total,
            entry.average_homework,
            entry.average_control,
            entry.term_grade,
        )
        for index, value in enumerate(values):
            total_cell = sheet.cell(row=row_index, column=totals_start + index)
            if value is not None:
                total_cell.value = value
                # Средние — до десятых, итог за четверть — целый.
                total_cell.number_format = "0" if index == 3 else "0.0"
            total_cell.alignment = CENTER
            total_cell.border = BORDER
            total_cell.fill = TOTAL_FILL
            if index == 3:
                total_cell.font = Font(bold=True)

    sheet.column_dimensions["A"].width = NAME_COLUMN_WIDTH
    sheet.freeze_panes = sheet.cell(row=first_data_row, column=2)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def file_name(group: StudyGroup, generated_at: datetime | None = None) -> str:
    # Название группы транслитерируем тем же правилом, что и логины: иначе
    # кириллица выпадет и выгрузки групп «7А» и «7Б» придут под одним именем.
    moment = generated_at or datetime.now(UTC)
    slug = re.sub(r"[^a-z0-9]+", "-", transliterate(group.name)).strip("-")
    return f"journal-{slug or 'group'}-{moment.strftime('%Y%m%d')}.xlsx"
